import json
from openpyxl.styles import Font
import openpyxl
from openpyxl.utils import get_column_letter

old_file = ""  #input("Enter old file path and name : ").strip() + ".xlsx"
new_file = ""  #input("Enter old file path and name : ").strip() + ".xlsx"

old_workbook = openpyxl.load_workbook(old_file)
new_workbook = openpyxl.load_workbook(new_file)

old_sheets = old_workbook.sheetnames
new_sheets = new_workbook.sheetnames

for a, b in zip(old_sheets, new_sheets):
    if a != b:
        print("Order of Sheets in both workbooks are not same...\nPlease correct the order and try again...")
        exit()

records = []

for a, b in zip(old_sheets, new_sheets):
    curr_old_sheet = old_workbook[a]
    curr_new_sheet = new_workbook[b]
    if curr_new_sheet.sheet_state != "visible":
        continue
    column_number = 4
    while True:
        row_number = 1
        while row_number < 1000:
            old_cell_value = curr_old_sheet.cell(row=row_number, column=column_number).value
            new_cell_value = curr_new_sheet.cell(row=row_number, column=column_number).value
            if str(old_cell_value).strip() != str(new_cell_value).strip():
                records.append({a: get_column_letter(column_number) + str(row_number) + str( old_cell_value)+"->"+str( new_cell_value)})
            row_number += 1

        old_cell_value = curr_old_sheet.cell(row=1, column=column_number).value
        new_cell_value = curr_new_sheet.cell(row=1, column=column_number).value

        if (old_cell_value is None or old_cell_value == "") and (new_cell_value is None or new_cell_value == ""):
            break

        column_number += 1

f = open("Differences.json", "w")
json.dump(records, f)
f.close()